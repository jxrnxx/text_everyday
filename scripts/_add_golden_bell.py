# -*- coding: utf-8 -*-
"""Add ability_public_golden_bell to 技能表.xlsx and item_book_golden_bell_1 to 物品表.xlsx"""
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ===== 1. 技能表 =====
ABILITY_EXCEL = os.path.join(BASE_DIR, 'excels', '技能表.xlsx')
wb_ability = openpyxl.load_workbook(ABILITY_EXCEL)
ws_ability = wb_ability.active

# Check if already exists
exists = False
for row in range(3, ws_ability.max_row + 1):
    if ws_ability.cell(row=row, column=1).value == 'ability_public_golden_bell':
        exists = True
        print(f"ability_public_golden_bell already exists at row {row}")
        break

if not exists:
    new_row = ws_ability.max_row + 1
    data = {
        1: 'ability_public_golden_bell',            # name
        2: 'ability_lua',                            # BaseClass
        3: 'abilities/ability_public_golden_bell',   # ScriptFile
        4: 'golden_bell_shield',                       # AbilityTextureName
        5: 'DOTA_ABILITY_BEHAVIOR_NO_TARGET',        # AbilityBehavior
        6: 4,                                         # MaxLevel
        7: 8,                                         # Cooldown
        8: 10,                                        # ManaCost
        15: 'PUBLIC',                                  # AbilityCategory
        16: 1,                                         # AbilityStar
        17: 'GENERAL',                                 # AbilityElement (通用)
        19: 'dmg_multiplier 5 7 9 11',                 # Value1
        20: 'shield_multiplier 5',                     # Value2
        21: 'shield_duration 3',                       # Value3
        22: 'explosion_radius 400',                    # Value4
        31: 'particles/units/heroes/hero_omniknight/omniknight_repel_buff.vpcf', # particle
        34: '通用 · 金钟罩',                            # Name CN
        35: '主动：运气震碎体表护盾，对周围造成伤害。护盾：根骨×5，持续3秒；伤害：根骨×倍率。',  # Desc CN
    }
    for col, val in data.items():
        ws_ability.cell(new_row, col, val)
    wb_ability.save(ABILITY_EXCEL)
    print(f"Added ability_public_golden_bell at row {new_row}")
else:
    print("Skipping ability add")

# ===== 2. 物品表 =====
ITEM_EXCEL = os.path.join(BASE_DIR, 'excels', '物品表.xlsx')
wb_item = openpyxl.load_workbook(ITEM_EXCEL)
ws_item = wb_item['npc_items_custom']

# Read headers
headers = {}
for c in range(1, ws_item.max_column + 1):
    h = ws_item.cell(row=2, column=c).value
    if h:
        headers[str(h).strip()] = c

# Check if already exists
exists = False
for row in range(3, ws_item.max_row + 1):
    if ws_item.cell(row=row, column=1).value == 'item_book_golden_bell_1':
        exists = True
        print(f"item_book_golden_bell_1 already exists at row {row}")
        break

if not exists:
    # Copy martial_cleave structure
    ref_row = None
    for row in range(3, ws_item.max_row + 1):
        if ws_item.cell(row=row, column=1).value == 'item_book_martial_cleave_1':
            ref_row = row
            break

    if ref_row:
        new_row = ws_item.max_row + 1
        for c in range(1, ws_item.max_column + 1):
            val = ws_item.cell(row=ref_row, column=c).value
            if val is not None:
                ws_item.cell(row=new_row, column=c, value=val)

        # Override fields
        ws_item.cell(row=new_row, column=1, value='item_book_golden_bell_1')
        if '#LocItemCn_{}' in headers:
            ws_item.cell(row=new_row, column=headers['#LocItemCn_{}'], value='金钟罩 技能书')
        if '#LocItemDesc_{}' in headers:
            ws_item.cell(row=new_row, column=headers['#LocItemDesc_{}'],
                value='学习后获得<font color="#ffcc66">主动技能</font>：激活<font color="#f5d442">金钟罩</font>，获得护盾并在结束时<font color="#ff6666">爆炸伤害</font>周围敌人')
        if 'AbilityTextureName' in headers:
            ws_item.cell(row=new_row, column=headers['AbilityTextureName'], value='golden_bell_shield')
        if 'ID' in headers:
            ws_item.cell(row=new_row, column=headers['ID'], value=1403)
        if 'LearnAbilityName' in headers:
            ws_item.cell(row=new_row, column=headers['LearnAbilityName'], value='ability_public_golden_bell')
        if 'IconPath' in headers:
            ws_item.cell(row=new_row, column=headers['IconPath'], value='file://{images}/spellicons/golden_bell_shield.png')
        if 'SkillBookCategory' in headers:
            ws_item.cell(row=new_row, column=headers['SkillBookCategory'], value='GENERAL')

        wb_item.save(ITEM_EXCEL)
        print(f"Added item_book_golden_bell_1 at row {new_row}")
    else:
        print("ERROR: Cannot find martial_cleave reference row")
else:
    print("Skipping item add")

print("\nDone! Run gen scripts to regenerate.")
