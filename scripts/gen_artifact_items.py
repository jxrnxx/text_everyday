"""
从 excels/物品表.xlsx 生成:
1. npc_items_artifacts.txt (DOTAItems 格式)
2. npc_items_custom.txt (DOTAItems 格式, 含 #base 引用)
3. 追加本地化条目到 addon.csv (中英文物品名)

用法: python scripts/gen_artifact_items.py
"""
import openpyxl
import os
import csv

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'excels', '物品表.xlsx')
NPC_DIR = os.path.join(BASE_DIR, 'game', 'scripts', 'npc')
ADDON_CSV = os.path.join(BASE_DIR, 'game', 'resource', 'addon.csv')

def read_sheet(wb, sheet_name):
    """读取 sheet，返回 [(item_name, {field: value}), ...]"""
    if sheet_name not in wb.sheetnames:
        print(f'WARNING: Sheet "{sheet_name}" not found')
        return []

    ws = wb[sheet_name]
    # Row 2 = KV field names
    headers = {}
    for c in range(1, ws.max_column + 1):
        key = ws.cell(row=2, column=c).value
        if key:
            headers[c] = str(key).strip()

    if not headers:
        return []

    name_col = 1
    items = []
    for row in range(3, ws.max_row + 1):
        item_name = ws.cell(row=row, column=name_col).value
        if not item_name:
            continue
        item_name = str(item_name).strip()
        if not item_name:
            continue

        fields = {}
        for c, key in headers.items():
            if c == name_col:
                continue
            val = ws.cell(row=row, column=c).value
            if val is not None and str(val).strip() != '':
                fields[key] = str(val).strip()
        items.append((item_name, fields))

    return items


def generate_kv_file(items, output_path, header_comment, base_includes=None):
    """生成 DOTAItems KV 文件"""
    lines = []
    lines.append(f'// this file is auto-generated from 物品表.xlsx')
    lines.append(f'// {header_comment}')
    lines.append('// DO NOT EDIT MANUALLY - edit the Excel file instead')

    if base_includes:
        for inc in base_includes:
            lines.append(f'#base "{inc}"')
        lines.append('')

    lines.append('"DOTAItems"')
    lines.append('{')

    # DisplayName_EN 不写入 KV，DisplayName 写入为本地化 token
    exclude_fields = {'DisplayName_EN'}

    for item_name, fields in items:
        lines.append(f'    "{item_name}"')
        lines.append('    {')
        kv_fields = {}
        for k, v in fields.items():
            if k in exclude_fields:
                continue
            if k == 'DisplayName':
                # 中文显示名 → 写入本地化 token 引用
                kv_fields[k] = f'#DOTA_Tooltip_ability_{item_name}'
            else:
                kv_fields[k] = v
        if kv_fields:
            max_key_len = max(len(k) for k in kv_fields.keys())
            for key, val in kv_fields.items():
                padding = ' ' * (max_key_len - len(key) + 4)
                lines.append(f'        "{key}"{padding}"{val}"')
        lines.append('    }')
        lines.append('')

    lines.append('}')
    lines.append('')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f'Generated {output_path} ({len(items)} items)')


def update_localization(all_items):
    """追加物品本地化条目到 addon.csv"""
    # 读取现有 CSV
    existing_tokens = set()
    existing_rows = []
    if os.path.exists(ADDON_CSV):
        with open(ADDON_CSV, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                existing_rows.append(row)
                if row:
                    existing_tokens.add(row[0])

    # 收集所有物品的本地化条目
    item_tokens = {}  # token -> [en_name, cn_name]
    for item_name, fields in all_items:
        cn_name = fields.get('DisplayName', '')
        en_name = fields.get('DisplayName_EN', item_name)
        if cn_name:
            token = f'DOTA_Tooltip_ability_{item_name}'
            item_tokens[token] = [en_name, cn_name]

    if not item_tokens:
        print('No localization entries needed')
        return

    # 更新现有 CSV: 覆盖已有的、追加新的
    updated_rows = []
    seen_tokens = set()
    if os.path.exists(ADDON_CSV):
        with open(ADDON_CSV, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0] in item_tokens:
                    # 覆盖已有条目
                    token = row[0]
                    updated_rows.append([token, item_tokens[token][0], item_tokens[token][1]])
                    seen_tokens.add(token)
                else:
                    updated_rows.append(row)

    # 追加新条目
    new_count = 0
    for token, (en_name, cn_name) in item_tokens.items():
        if token not in seen_tokens:
            updated_rows.append([token, en_name, cn_name])
            new_count += 1

    with open(ADDON_CSV, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(updated_rows)
    print(f'Localization: {len(seen_tokens)} updated, {new_count} added to {ADDON_CSV}')


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # 1. 生成 npc_items_artifacts.txt
    artifact_items = read_sheet(wb, 'npc_items_artifacts')
    if artifact_items:
        generate_kv_file(
            artifact_items,
            os.path.join(NPC_DIR, 'npc_items_artifacts.txt'),
            'Sheet: npc_items_artifacts (装备神器)'
        )

    # 2. 生成 npc_items_custom.txt
    custom_items = read_sheet(wb, 'npc_items_custom')
    if custom_items:
        generate_kv_file(
            custom_items,
            os.path.join(NPC_DIR, 'npc_items_custom.txt'),
            'Sheet: npc_items_custom (通用物品)',
            base_includes=['npc_items_artifacts.txt']
        )

    # 3. 更新本地化
    all_items = artifact_items + custom_items
    update_localization(all_items)

    print('\nDone!')


if __name__ == '__main__':
    main()
