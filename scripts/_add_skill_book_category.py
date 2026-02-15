# -*- coding: utf-8 -*-
"""Add SkillBookCategory column and fix names in 物品表.xlsx npc_items_custom sheet"""
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'excels', '物品表.xlsx')

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['npc_items_custom']

# Read headers
headers = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=2, column=c).value
    if h:
        headers[str(h).strip()] = c

print("Current headers:", headers)

# Check if SkillBookCategory column already exists
if 'SkillBookCategory' in headers:
    cat_col = headers['SkillBookCategory']
    print(f"SkillBookCategory already exists at col {cat_col}")
else:
    # Add new column after IsSkillBook
    cat_col = ws.max_column + 1
    ws.cell(row=1, column=cat_col, value='技能书分类')  # Chinese header
    ws.cell(row=2, column=cat_col, value='SkillBookCategory')  # English KV key
    print(f"Added SkillBookCategory at col {cat_col}")

# Find and update skill book rows
name_col = headers.get('#LocItemCn_{}') or 2  # display name column

for row in range(3, ws.max_row + 1):
    item_name = ws.cell(row=row, column=1).value
    if not item_name:
        continue

    if item_name == 'item_book_martial_cleave_1':
        ws.cell(row=row, column=name_col, value='横扫 技能书')
        ws.cell(row=row, column=cat_col, value='MARTIAL')
        print(f"  Row {row}: {item_name} -> name='横扫 技能书', cat=MARTIAL")

    elif item_name == 'item_book_plague_cloud_1':
        ws.cell(row=row, column=name_col, value='噬魂毒阵 技能书')
        ws.cell(row=row, column=cat_col, value='DIVINITY')
        # Also fix desc
        desc_col = headers.get('#LocItemDesc_{}') or 3
        ws.cell(row=row, column=desc_col, value='学习后获得<font color="#ffcc66">主动技能</font>：在目标区域布下<font color="#66ff66">噬魂毒阵</font>，阵内敌人每秒受到<font color="#66ccff">法术伤害</font>')
        # Fix icon
        icon_col = headers.get('IconPath')
        if icon_col:
            ws.cell(row=row, column=icon_col, value='file://{images}/spellicons/plague_cloud_icon.png')
        print(f"  Row {row}: {item_name} -> name='噬魂毒阵 技能书', cat=DIVINITY")

wb.save(EXCEL_PATH)
print("\nExcel updated with SkillBookCategory!")
