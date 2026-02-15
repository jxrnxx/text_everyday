# -*- coding: utf-8 -*-
"""
向 物品表.xlsx 的 npc_items_custom sheet 添加 item_book_plague_cloud_1
然后运行 gen_artifact_items.py 重新生成
"""
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'excels', '物品表.xlsx')

wb = openpyxl.load_workbook(EXCEL_PATH)

sheet_name = 'npc_items_custom'
if sheet_name not in wb.sheetnames:
    print(f'ERROR: Sheet "{sheet_name}" not found in {EXCEL_PATH}')
    print(f'Available sheets: {wb.sheetnames}')
    exit(1)

ws = wb[sheet_name]

# 先读取 row 2 的 header 映射，了解每列什么字段
print("=== Sheet Headers (Row 2) ===")
headers = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=2, column=c).value
    if h:
        headers[str(h).strip()] = c
        print(f"  Col {c}: {h}")

print(f"\n=== Current Items (Row 3+) ===")
for row in range(3, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name:
        print(f"  Row {row}: {name}")

# 检查是否已存在
target = 'item_book_plague_cloud_1'
exists = False
for row in range(3, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == target:
        exists = True
        print(f"\n{target} already exists at row {row}")
        break

if not exists:
    # 找 item_book_martial_cleave_1 作为参考
    ref_row = None
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 'item_book_martial_cleave_1':
            ref_row = row
            break

    if ref_row:
        print(f"\nReference: item_book_martial_cleave_1 at row {ref_row}")
        print("  Fields:")
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=ref_row, column=c).value
            if val is not None:
                h = ws.cell(row=2, column=c).value or f'Col{c}'
                print(f"    {h} (col {c}): {val}")

        # 复制 martial_cleave 的行，修改为 plague_cloud
        new_row = ws.max_row + 1
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=ref_row, column=c).value
            if val is not None:
                ws.cell(row=new_row, column=c, value=val)

        # 覆盖特定字段
        ws.cell(row=new_row, column=1, value='item_book_plague_cloud_1')

        # 修改 DisplayName
        if 'DisplayName' in headers:
            ws.cell(row=new_row, column=headers['DisplayName'], value='神念·噬魂毒阵 技能书')
        if 'DisplayName_EN' in headers:
            ws.cell(row=new_row, column=headers['DisplayName_EN'], value='Soul-Devouring Poison Formation Skill Book')

        # 修改 AbilityTextureName
        if 'AbilityTextureName' in headers:
            ws.cell(row=new_row, column=headers['AbilityTextureName'], value='item_book_plague_cloud_1')

        # 修改 LearnAbilityName (如果有的话)
        if 'LearnAbilityName' in headers:
            ws.cell(row=new_row, column=headers['LearnAbilityName'], value='ability_public_plague_cloud')

        # 修改 ID (如果有的话) - martial_cleave 是 1401，plague_cloud 用 1402
        if 'ID' in headers:
            ws.cell(row=new_row, column=headers['ID'], value=1402)

        # 修改 ScriptFile (如果有的话)
        if 'ScriptFile' in headers:
            ws.cell(row=new_row, column=headers['ScriptFile'], value='items/item_learn_skill')

        wb.save(EXCEL_PATH)
        print(f"\nDone! Added {target} at row {new_row}")
    else:
        print("\nERROR: Could not find item_book_martial_cleave_1 as reference")
else:
    print("Skipping add - already exists")

print("\n=== Now run gen_artifact_items.py to regenerate ===")
