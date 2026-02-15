# -*- coding: utf-8 -*-
"""Fix item_book_plague_cloud_1 fields in 物品表.xlsx"""
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'excels', '物品表.xlsx')

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['npc_items_custom']

# Find headers
headers = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(row=2, column=c).value
    if h:
        headers[str(h).strip()] = c

# Find plague_cloud row
target_row = None
for row in range(3, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == 'item_book_plague_cloud_1':
        target_row = row
        break

if not target_row:
    print("ERROR: item_book_plague_cloud_1 not found!")
    exit(1)

print(f"Found item_book_plague_cloud_1 at row {target_row}")

# Update fields
updates = {
    '#LocItemCn_{}': '神念·噬魂毒阵 技能书',
    '#LocItemDesc_{}': '学习后获得<font color="#ffcc66">主动技能</font>：在目标区域布下<font color="#66ff66">噬魂毒阵</font>，阵内敌人每秒受到<font color="#66ccff">法术伤害</font>',
    'IconPath': 'file://{images}/spellicons/plague_cloud_icon.png',
    'LearnAbilityName': 'ability_public_plague_cloud',
}

for field_name, new_value in updates.items():
    if field_name in headers:
        col = headers[field_name]
        old_val = ws.cell(row=target_row, column=col).value
        ws.cell(row=target_row, column=col, value=new_value)
        print(f"  {field_name} (col {col}): '{old_val}' -> '{new_value}'")
    else:
        print(f"  WARNING: field '{field_name}' not found in headers")

wb.save(EXCEL_PATH)
print("\nExcel updated! Now regenerating...")
