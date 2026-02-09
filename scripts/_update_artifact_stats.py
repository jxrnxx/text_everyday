"""
更新 物品表.xlsx 中的 npc_items_artifacts sheet
按用户提供的新数值表更新所有装备属性

新属性表:
| 槽位 | 属性       | T0     | T1     | T2        | T3         | T4          | T5           |
|------|-----------|--------|--------|-----------|------------|-------------|--------------|
| 武器 | 攻/穿     | 15/-   | 60/-   | 300/10    | 1500/40    | 8000/150    | 50000/500    |
| 衣服 | 根骨/甲   | 5/-    | 20/2   | 100/8     | 500/20     | 3000/40     | 20000/60     |
| 头盔 | 技伤/格挡  | 2/10   | 10/30  | 25/80     | 50/5%(减伤)| 100/10%     | 200/20%      |
| 饰品 | 暴击/爆伤  | 2/-    | 5/-    | 10/-      | 15/30      | 20/80       | 25/200       |
| 鞋子 | 身法/闪避  | 5/-    | 15/-   | 80/-      | 300/5      | 2000/15     | 10000/25     |
| 护符 | 全属/终伤  | 2/-    | 10/-   | 40/-      | 150/2      | 1000/5      | 5000/10      |
"""
import openpyxl
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'excels', '物品表.xlsx')

# ============================================================
# 新属性数据定义
# ============================================================

# 每个条目: item_name -> {KV_field: value}
# 只列出需要变更的 stat 字段 (BaseClass / ScriptFile 等不动)
ARTIFACT_STATS = {
    # ===== 武器 (Slot 0) =====
    'item_artifact_weapon_t0': {
        'BonusDamage': 15,
    },
    'item_artifact_weapon_tier1': {
        'BonusDamage': 60,
    },
    'item_artifact_weapon_t2': {
        'BonusDamage': 300, 'BonusArmorPen': 10,
    },
    'item_artifact_weapon_t3': {
        'BonusDamage': 1500, 'BonusArmorPen': 40,
    },
    'item_artifact_weapon_t4': {
        'BonusDamage': 8000, 'BonusArmorPen': 150,
    },
    'item_artifact_weapon_t5': {
        'BonusDamage': 50000, 'BonusArmorPen': 500,
    },

    # ===== 衣甲 (Slot 1): 根骨 + 甲 =====
    'item_artifact_armor_t0': {
        'BonusConstitution': 5,
    },
    'item_artifact_armor_tier1': {
        'BonusConstitution': 20, 'BonusArmor': 2,
    },
    'item_artifact_armor_t2': {
        'BonusConstitution': 100, 'BonusArmor': 8,
    },
    'item_artifact_armor_t3': {
        'BonusConstitution': 500, 'BonusArmor': 20,
    },
    'item_artifact_armor_t4': {
        'BonusConstitution': 3000, 'BonusArmor': 40,
    },
    'item_artifact_armor_t5': {
        'BonusConstitution': 20000, 'BonusArmor': 60,
    },

    # ===== 头盔 (Slot 2): 技伤% + 格挡(T0-T2) / 减伤%(T3-T5) =====
    'item_artifact_helm_t0': {
        'BonusSpellDamage': 2, 'BonusBlock': 10,
    },
    'item_artifact_helm_tier1': {
        'BonusSpellDamage': 10, 'BonusBlock': 30,
    },
    'item_artifact_helm_t2': {
        'BonusSpellDamage': 25, 'BonusBlock': 80,
    },
    'item_artifact_helm_t3': {
        'BonusSpellDamage': 50, 'BonusFinalDmgReduct': 5,
    },
    'item_artifact_helm_t4': {
        'BonusSpellDamage': 100, 'BonusFinalDmgReduct': 10,
    },
    'item_artifact_helm_t5': {
        'BonusSpellDamage': 200, 'BonusFinalDmgReduct': 20,
    },

    # ===== 饰品 (Slot 3): 暴击% + 爆伤% =====
    'item_artifact_accessory_t0': {
        'BonusCritChance': 2,
    },
    'item_artifact_accessory_tier1': {
        'BonusCritChance': 5,
    },
    'item_artifact_accessory_t2': {
        'BonusCritChance': 10,
    },
    'item_artifact_accessory_t3': {
        'BonusCritChance': 15, 'BonusCritDamage': 30,
    },
    'item_artifact_accessory_t4': {
        'BonusCritChance': 20, 'BonusCritDamage': 80,
    },
    'item_artifact_accessory_t5': {
        'BonusCritChance': 25, 'BonusCritDamage': 200,
    },

    # ===== 鞋子 (Slot 4): 身法 + 闪避% =====
    'item_artifact_boots_t0': {
        'BonusAgility': 5,
    },
    'item_artifact_boots_tier1': {
        'BonusAgility': 15,
    },
    'item_artifact_boots_t2': {
        'BonusAgility': 80,
    },
    'item_artifact_boots_t3': {
        'BonusAgility': 300, 'BonusEvasion': 5,
    },
    'item_artifact_boots_t4': {
        'BonusAgility': 2000, 'BonusEvasion': 15,
    },
    'item_artifact_boots_t5': {
        'BonusAgility': 10000, 'BonusEvasion': 25,
    },

    # ===== 护符 (Slot 5): 全属性 + 终伤增% =====
    'item_artifact_amulet_t0': {
        'BonusAllStats': 2,
    },
    'item_artifact_amulet_tier1': {
        'BonusAllStats': 10,
    },
    'item_artifact_amulet_t2': {
        'BonusAllStats': 40,
    },
    'item_artifact_amulet_t3': {
        'BonusAllStats': 150, 'BonusFinalDmgIncrease': 2,
    },
    'item_artifact_amulet_t4': {
        'BonusAllStats': 1000, 'BonusFinalDmgIncrease': 5,
    },
    'item_artifact_amulet_t5': {
        'BonusAllStats': 5000, 'BonusFinalDmgIncrease': 10,
    },
}

# 需要清除的旧字段 (按物品名分组)
# 衣甲: 移除 BonusHP (改为 BonusConstitution)
# 头盔: 移除 BonusDivinity, BonusManaRegen (改为 BonusSpellDamage + BonusBlock/FinalDmgReduct)
# 鞋子: 移除 BonusMoveSpeed (改为 BonusAgility)
# 护符: 移除 BonusFinalDmgReduct (T3-T5, 改为 BonusFinalDmgIncrease)
FIELDS_TO_CLEAR = {
    # 衣甲 - 移除 BonusHP
    'item_artifact_armor_t0': ['BonusHP'],
    'item_artifact_armor_tier1': ['BonusHP'],
    'item_artifact_armor_t2': ['BonusHP'],
    'item_artifact_armor_t3': ['BonusHP'],
    'item_artifact_armor_t4': ['BonusHP'],
    'item_artifact_armor_t5': ['BonusHP'],
    # 头盔 - 移除 BonusDivinity, BonusManaRegen
    'item_artifact_helm_t0': ['BonusDivinity', 'BonusManaRegen'],
    'item_artifact_helm_tier1': ['BonusDivinity', 'BonusManaRegen'],
    'item_artifact_helm_t2': ['BonusDivinity', 'BonusManaRegen'],
    'item_artifact_helm_t3': ['BonusDivinity', 'BonusManaRegen'],
    'item_artifact_helm_t4': ['BonusDivinity', 'BonusManaRegen'],
    'item_artifact_helm_t5': ['BonusDivinity', 'BonusManaRegen'],
    # 鞋子 - 移除 BonusMoveSpeed
    'item_artifact_boots_t0': ['BonusMoveSpeed'],
    'item_artifact_boots_tier1': ['BonusMoveSpeed'],
    'item_artifact_boots_t2': ['BonusMoveSpeed'],
    'item_artifact_boots_t3': ['BonusMoveSpeed'],
    'item_artifact_boots_t4': ['BonusMoveSpeed'],
    'item_artifact_boots_t5': ['BonusMoveSpeed'],
    # 护符 - 移除 BonusFinalDmgReduct (替换为 BonusFinalDmgIncrease)
    'item_artifact_amulet_t3': ['BonusFinalDmgReduct'],
    'item_artifact_amulet_t4': ['BonusFinalDmgReduct'],
    'item_artifact_amulet_t5': ['BonusFinalDmgReduct'],
}


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f'ERROR: Excel not found: {EXCEL_PATH}')
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet_name = 'npc_items_artifacts'

    if sheet_name not in wb.sheetnames:
        print(f'ERROR: Sheet "{sheet_name}" not found')
        sys.exit(1)

    ws = wb[sheet_name]

    # 1. 读取表头 (Row 2 = KV field names)
    headers = {}  # col_index -> field_name
    field_to_col = {}  # field_name -> col_index
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        key = ws.cell(row=2, column=c).value
        if key:
            key = str(key).strip()
            headers[c] = key
            field_to_col[key] = c

    print(f'当前表头: {list(field_to_col.keys())}')

    # 2. 确保新字段列存在
    new_fields = ['BonusSpellDamage', 'BonusBlock', 'BonusFinalDmgIncrease',
                  'BonusConstitution', 'BonusAgility']
    for field in new_fields:
        if field not in field_to_col:
            max_col += 1
            ws.cell(row=1, column=max_col).value = {
                'BonusSpellDamage': '技能伤害%',
                'BonusBlock': '格挡',
                'BonusFinalDmgIncrease': '终伤增%',
                'BonusConstitution': '根骨',
                'BonusAgility': '身法',
            }.get(field, field)
            ws.cell(row=2, column=max_col).value = field
            field_to_col[field] = max_col
            headers[max_col] = field
            print(f'  新增列 {max_col}: {field}')

    # 3. 建立 item_name -> row 映射
    item_rows = {}  # item_name -> row_number
    for row in range(3, ws.max_row + 1):
        item_name = ws.cell(row=row, column=1).value
        if item_name:
            item_name = str(item_name).strip()
            item_rows[item_name] = row

    print(f'找到 {len(item_rows)} 个物品')

    # 4. 清除旧字段
    cleared = 0
    for item_name, fields_list in FIELDS_TO_CLEAR.items():
        row = item_rows.get(item_name)
        if not row:
            print(f'  WARNING: 找不到 {item_name}')
            continue
        for field in fields_list:
            col = field_to_col.get(field)
            if col:
                old_val = ws.cell(row=row, column=col).value
                if old_val is not None:
                    ws.cell(row=row, column=col).value = None
                    cleared += 1

    print(f'清除了 {cleared} 个旧字段值')

    # 5. 写入新数据
    updated = 0
    for item_name, stats in ARTIFACT_STATS.items():
        row = item_rows.get(item_name)
        if not row:
            print(f'  WARNING: 找不到 {item_name}')
            continue
        for field, value in stats.items():
            col = field_to_col.get(field)
            if not col:
                print(f'  WARNING: 找不到字段 {field}')
                continue
            ws.cell(row=row, column=col).value = value
            updated += 1

    print(f'更新了 {updated} 个字段值')

    # 6. 保存
    wb.save(EXCEL_PATH)
    print(f'\n已保存到 {EXCEL_PATH}')
    print('请运行 python scripts/gen_artifact_items.py 重新生成 KV 文件')


if __name__ == '__main__':
    main()
