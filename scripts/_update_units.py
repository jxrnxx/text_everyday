"""
更新 单位表.xlsx 中的 custom_units sheet
按用户提供的新数值表更新所有单位属性

Excel 列结构 (Row2 KV字段名):
Col1=UnitName, Col2=#LocUnitNameCn_{}, Col4=BaseClass, Col5=Model,
Col7=ModelScale, Col33=HealthBarOffset, Col34=MovementSpeed,
Col37=MovementCapabilities, Col38=AttackCapabilities, Col39=ArmorPhysical,
Col41=StatusHealth, Col42=StatusHealthRegen, Col43=StatusMana,
Col45=BountyGoldMin, Col46=BountyGoldMax, Col47=AttackDamageMin,
Col48=AttackDamageMax, Col49=AttackRate, Col50=AttackAnimationPoint,
Col51=AttackRange, Col52=BoundsHullName, Col55=AttackAcquisitionRange,
Col56=ChaseDistance, Col57=CombatClassAttack, Col58=CombatClassDefend,
Col59=CustomDrop_Coin, Col60=CustomDrop_Faith, Col61=StatLabel,
Col62=HaveLevel, Col63=CustomDrop_DefenderPoints,
Col64=Artifact_Drop_Type, Col65=Artifact_Drop_XP
"""
import openpyxl
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'excels', '单位表.xlsx')

# ============================================================
# 行更新数据：(UnitName, 中文名, 数据字典)
# 数据字典用 Row2 的 KV 字段名做 key
# ============================================================

def mk(base, model, scale, hbo, spd, movCap, atkCap, armor, hp, hpReg, mana,
       goldMin=None, goldMax=None, atkMin=None, atkMax=None,
       atkRate=None, atkAnim=None, atkRange=None, hull=None,
       acqRange=None, chase=None, combAtk=None, combDef=None,
       coin=None, faith=None, stat=None, exp=None, defPts=None,
       artType=None, artXP=None):
    d = {
        'BaseClass': base, 'Model': model, 'ModelScale': scale,
        'HealthBarOffset': hbo, 'MovementCapabilities': movCap,
        'AttackCapabilities': atkCap, 'ArmorPhysical': armor,
        'StatusHealth': hp, 'StatusHealthRegen': hpReg, 'StatusMana': mana,
    }
    if spd is not None: d['MovementSpeed'] = spd
    if goldMin is not None: d['BountyGoldMin'] = goldMin
    if goldMax is not None: d['BountyGoldMax'] = goldMax
    if atkMin is not None: d['AttackDamageMin'] = atkMin
    if atkMax is not None: d['AttackDamageMax'] = atkMax
    if atkRate is not None: d['AttackRate'] = atkRate
    if atkAnim is not None: d['AttackAnimationPoint'] = atkAnim
    if atkRange is not None: d['AttackRange'] = atkRange
    if hull is not None: d['BoundsHullName'] = hull
    if acqRange is not None: d['AttackAcquisitionRange'] = acqRange
    if chase is not None: d['ChaseDistance'] = chase
    if combAtk is not None: d['CombatClassAttack'] = combAtk
    if combDef is not None: d['CombatClassDefend'] = combDef
    if coin is not None: d['CustomDrop_Coin'] = coin
    if faith is not None: d['CustomDrop_Faith'] = faith
    if stat is not None: d['StatLabel'] = stat
    if exp is not None: d['HaveLevel'] = exp
    if defPts is not None: d['CustomDrop_DefenderPoints'] = defPts
    if artType is not None: d['Artifact_Drop_Type'] = artType
    if artXP is not None: d['Artifact_Drop_XP'] = artXP
    return d

CB = 'DOTA_COMBAT_CLASS_ATTACK_BASIC'
CD = 'DOTA_COMBAT_CLASS_DEFEND_BASIC'
MG = 'DOTA_UNIT_CAP_MOVE_GROUND'
MN = 'DOTA_UNIT_CAP_MOVE_NONE'
AM = 'DOTA_UNIT_CAP_MELEE_ATTACK'
AN = 'DOTA_UNIT_CAP_NO_ATTACK'
HH = 'DOTA_HULL_SIZE_HERO'
HR = 'DOTA_HULL_SIZE_REGULAR'
HB = 'DOTA_HULL_SIZE_BUILDING'

# 重命名映射：旧名 -> 新名
RENAME_MAP = {
    'npc_enemy_zombie_lvl1': 'npc_creep_train_tier1',
    'npc_enemy_zombie_lvl2': 'npc_creep_train_tier2',
    'npc_enemy_zombie_lvl3': 'npc_creep_train_tier3',
    'npc_enemy_zombie_lvl4': 'npc_creep_train_tier4',
    'npc_enemy_zombie_lvl5': 'npc_creep_train_tier5',
    'npc_enemy_zombie_lvl6': 'npc_creep_train_tier6',
    'npc_enemy_zombie_lvl7': 'npc_creep_train_tier7',
    'npc_enemy_zombie_lvl8': 'npc_creep_train_tier8',
}

# 中文名映射
CN_NAMES = {
    'npc_dota_home_base': '一线希望',
    'npc_creep_train_tier1': '一阶·游荡梦魇',
    'npc_creep_train_tier2': '二阶·血肉傀儡',
    'npc_creep_train_tier3': '三阶·禁忌信徒',
    'npc_creep_train_tier4': '四阶·虚空行者',
    'npc_creep_train_tier5': '五阶·噩梦骑士',
    'npc_creep_train_tier6': '六阶·古神眷属',
    'npc_creep_train_tier7': '七阶·混沌衍生物',
    'npc_creep_train_tier8': '八阶·神孽',
    'npc_cultivation_merchant': '修炼商人',
    'npc_ability_merchant': '技能商人',
    'npc_creep_wave_1': '腐化猎犬',
    'npc_creep_wave_2': '狂暴僵尸',
    'npc_creep_wave_3': '骷髅射手',
    'npc_creep_wave_4': '黑暗卫士',
    'npc_creep_wave_5': '兵-血腥',
    'npc_creep_wave_6': '熔岩傀儡',
    'npc_creep_wave_7': '剧毒飞龙',
    'npc_creep_wave_8': '虚空行者',
    'npc_creep_wave_9': '禁忌信徒',
    'npc_creep_wave_10': '远古雷兽',
    'npc_creep_wave_11': '深渊骑士',
    'npc_creep_wave_12': '堕落神官',
    'npc_creep_wave_13': '巨型憎恶',
    'npc_creep_wave_14': '混沌元素',
    'npc_creep_wave_15': '兵-虚空恐',
    'npc_creep_wave_16': '灭世魔龙',
    'npc_creep_wave_17': '泰坦巨人',
    'npc_creep_wave_18': '虚空领主',
    'npc_creep_wave_19': '神之倒影',
    'npc_boss_wave_5': '血肉巨像',
    'npc_boss_wave_10': '深渊领主',
    'npc_boss_wave_15': '堕落神使',
    'npc_boss_wave_20': '戏神',
    'npc_guardian_zone_1': '剑冢守卫',
    'npc_guardian_zone_2': '枯木铁卫',
    'npc_guardian_zone_3': '熔火精灵',
    'npc_guardian_zone_4': '寒冰傀儡',
    'npc_guardian_zone_5': '雷行兽',
    'npc_guardian_zone_6': '遗迹守灵',
}

# 完整数据: UnitName -> 字段数据
UNITS = {}

# --- Home Base ---
UNITS['npc_dota_home_base'] = mk(
    'npc_dota_building', 'models/props_structures/radiant_tower002.vmdl', 1.3,
    400, None, MN, AN, 20, 20000, 50, 0, hull=HB, faith=0)

# --- 练功房怪 (train tiers) ---
_train = [
    ('npc_creep_train_tier1', 'models/creeps/neutral_creeps/n_creep_ghost_b/n_creep_ghost_b.vmdl', 1, 0, 90, 0.5, 10, 10, 15, 1, 0, 30, 0),
    ('npc_creep_train_tier2', 'models/heroes/life_stealer/life_stealer.vmdl', 1, 3, 500, 5, 50, 50, 45, 1, 1, 180, 0),
    ('npc_creep_train_tier3', 'models/heroes/necrolyte/necrolyte.vmdl', 1, 5, 3000, 0.5, 200, 200, 120, 1, 2, 450, 0),
    ('npc_creep_train_tier4', 'models/heroes/void_spirit/void_spirit.vmdl', 1, 10, 15000, 0.5, 1000, 1000, 300, 1, 3, 900, 0),
    ('npc_creep_train_tier5', 'models/heroes/chaos_knight/chaos_knight.vmdl', 1, 20, 80000, 0.5, 5000, 5000, 800, 1, 3, 2000, 0),
    ('npc_creep_train_tier6', 'models/heroes/faceless_void/faceless_void.vmdl', 1, 40, 300000, 0.5, 20000, 20000, 2000, 1, 4, 5000, 0),
    ('npc_creep_train_tier7', 'models/heroes/enigma/enigma.vmdl', 1, 80, 1500000, 0.5, 100000, 100000, 5000, 1, 5, 12000, 0),
    ('npc_creep_train_tier8', 'models/heroes/shadow_fiend/shadow_fiend_arcana.vmdl', 1, 150, 5000000, 0.5, 500000, 500000, 12000, 1, 5, 30000, 0),
]
for name, mdl, sc, ar, hp, hpR, amin, amax, coin, faith, stat, exp, defPts in _train:
    UNITS[name] = mk('npc_dota_creature', mdl, sc, 160, 300, MG, AM,
        ar, hp, hpR, 0, 0, 0, amin, amax, 1.5, 0.3, 128, HH,
        800, 1000, CB, CD, coin, faith, stat, exp, defPts)

# --- 商人 ---
UNITS['npc_cultivation_merchant'] = mk('npc_dota_creature',
    'models/props_gameplay/shopkeeper_fountain/shopkeeper_fountain.vmdl', 1,
    0, 0, MN, AN, 99, 9999, 0, 0, 0, 0, 0, 0, 1, 0, 0, HH, 0, 0,
    coin=0, defPts=10)
UNITS['npc_ability_merchant'] = mk('npc_dota_creature',
    'models/props_gameplay/side_shop_keeper/side_shop_keeper.vmdl', 1.1,
    0, 0, MN, AN, 99, 9999, 0, 0, 0, 0, 0, 0, 1, 0, 0, HH, 0, 0,
    coin=0, defPts=10)

# --- 波次怪 ---
_wave = [
    ('npc_creep_wave_1', 'models/creeps/neutral_creeps/n_creep_gnoll/n_creep_gnoll_frost.vmdl', 1, 0, 200, 0, 25, 25, 100, 25, 5, 0, 45, 10),
    ('npc_creep_wave_2', 'models/heroes/undying/undying_minion.vmdl', 1, 1, 500, 1, 40, 40, 100, 35, 5, 0, 60, 10),
    ('npc_creep_wave_3', 'models/creeps/neutral_creeps/n_creep_troll_skeleton/n_creep_troll_skeleton.vmdl', 1, 2, 800, 2, 60, 60, 600, 50, 5, 0, 80, 10),
    ('npc_creep_wave_4', 'models/creeps/lane_creeps/creep_bad_melee/creep_bad_melee.vmdl', 1, 3, 1500, 5, 90, 90, 100, 70, 5, 0, 120, 10),
    ('npc_creep_wave_5', 'models/creeps/neutral_creeps/n_creep_satyr_hellcaller/n_creep_satyr_hellcaller.vmdl', 1.2, 5, 3000, 10, 150, 150, 100, 100, 5, 1, 200, 10),
    ('npc_creep_wave_6', 'models/heroes/warlock/warlock_demon.vmdl', 0.8, 8, 6000, 20, 220, 220, 100, 150, 5, 1, 300, 10),
    ('npc_creep_wave_7', 'models/creeps/neutral_creeps/n_creep_black_drake/n_creep_black_drake.vmdl', 1, 10, 10000, 50, 350, 350, 600, 200, 5, 1, 400, 10),
    ('npc_creep_wave_8', 'models/heroes/void_spirit/void_spirit.vmdl', 1, 15, 25000, 100, 500, 500, 150, 300, 5, 1, 600, 10),
    ('npc_creep_wave_9', 'models/heroes/necrolyte/necrolyte.vmdl', 1, 20, 50000, 200, 800, 800, 600, 450, 5, 1, 1000, 10),
    ('npc_creep_wave_10', 'models/creeps/neutral_creeps/n_creep_thunder_lizard/n_creep_thunder_lizard_big.vmdl', 1.3, 30, 100000, 500, 1500, 1500, 150, 600, 5, 2, 1500, 10),
    ('npc_creep_wave_11', 'models/heroes/abaddon/abaddon.vmdl', 1.2, 40, 200000, 1000, 3000, 3000, 150, 1000, 5, 2, 2500, 10),
    ('npc_creep_wave_12', 'models/heroes/oracle/oracle.vmdl', 1.2, 50, 400000, 2000, 6000, 6000, 600, 1500, 5, 2, 4000, 10),
    ('npc_creep_wave_13', 'models/heroes/pudge/pudge.vmdl', 1.3, 70, 800000, 5000, 12000, 12000, 150, 2500, 5, 2, 6000, 10),
    ('npc_creep_wave_14', 'models/heroes/enigma/enigma.vmdl', 1, 100, 1500000, 10000, 25000, 25000, 600, 4000, 5, 2, 10000, 10),
    ('npc_creep_wave_15', 'models/heroes/shadow_demon/shadow_demon.vmdl', 1.2, 150, 3000000, 20000, 50000, 50000, 400, 6000, 5, 3, 15000, 10),
    ('npc_creep_wave_16', 'models/creeps/neutral_creeps/n_creep_black_dragon/n_creep_black_dragon.vmdl', 1.5, 200, 8000000, 50000, 100000, 100000, 600, 10000, 5, 3, 25000, 10),
    ('npc_creep_wave_17', 'models/heroes/tiny/tiny_04/tiny_04.vmdl', 1.5, 300, 20000000, 100000, 200000, 200000, 150, 15000, 5, 3, 40000, 10),
    ('npc_creep_wave_18', 'models/heroes/faceless_void/faceless_void.vmdl', 1.5, 400, 60000000, 200000, 400000, 400000, 150, 25000, 5, 3, 60000, 10),
    ('npc_creep_wave_19', 'models/heroes/terrorblade/demon_form/terrorblade_demon_form.vmdl', 1.8, 500, 200000000, 500000, 800000, 800000, 150, 50000, 5, 4, 100000, 10),
]
for name, mdl, sc, ar, hp, hpR, amin, amax, atkRng, coin, faith, stat, exp, defPts in _wave:
    UNITS[name] = mk('npc_dota_creature', mdl, sc, 160, 300, MG, AM,
        ar, hp, hpR, 0, 0, 0, amin, amax, 1.5, 0.3, atkRng, HH,
        800, 1000, CB, CD, coin, faith, stat, exp, defPts)

# --- BOSS ---
_boss = [
    ('npc_boss_wave_5', 'models/heroes/life_stealer/life_stealer.vmdl', 1.8, 10, 20000, 100, 500, 500, 150, 2000, 50, 1, 5000, 200),
    ('npc_boss_wave_10', 'models/heroes/doom/doom.vmdl', 2, 30, 500000, 5000, 5000, 5000, 150, 5000, 50, 2, 50000, 500),
    ('npc_boss_wave_15', 'models/heroes/obsidian_destroyer/obsidian_destroyer.vmdl', 2.2, 100, 20000000, 100000, 200000, 200000, 600, 20000, 50, 3, 200000, 500),
    ('npc_boss_wave_20', 'models/heroes/grimstroke/grimstroke.vmdl', 3, 1000, 3000000000, 2000000, 5000000, 5000000, 800, 0, 100, 5, 1000000, 5000),
]
for name, mdl, sc, ar, hp, hpR, amin, amax, atkRng, coin, faith, stat, exp, defPts in _boss:
    UNITS[name] = mk('npc_dota_creature', mdl, sc, 160, 300, MG, AM,
        ar, hp, hpR, 0, 0, 0, amin, amax, 1.5, 0.3, atkRng, HH,
        800, 1000, CB, CD, coin, faith, stat, exp, defPts)

# --- 守卫 ---
_guard = [
    ('npc_guardian_zone_1', 'models/creeps/neutral_creeps/n_creep_ancient_frog/n_creep_ancient_frog_mage.vmdl', 1, 0, 300, 1, 22, 22, 150, 30, 1, 0, 50, 0, 1, 50),
    ('npc_guardian_zone_2', 'models/creeps/neutral_creeps/n_creep_froglet/n_creep_froglet_mage.vmdl', 0.8, 2, 400, 1, 15, 15, 150, 30, 1, 0, 50, 0, 2, 50),
    ('npc_guardian_zone_3', 'models/creeps/neutral_creeps/n_creep_ghost_b/n_creep_ghost_b.vmdl', 1, 0, 300, 1, 20, 20, 150, 30, 1, 0, 50, 0, 3, 50),
    ('npc_guardian_zone_4', 'models/items/broodmother/spiderling/the_glacial_creeper_creepling/the_glacial_creeper_creepling.vmdl', 1, 0, 300, 1, 20, 20, 150, 30, 1, 0, 50, 0, 4, 50),
    ('npc_guardian_zone_5', 'models/creeps/neutral_creeps/n_creep_tadpole_c/n_creep_tadpole_c.vmdl', 0.9, 0, 300, 1, 25, 25, 150, 30, 1, 0, 50, 0, 5, 50),
    ('npc_guardian_zone_6', 'models/creeps/lane_creeps/creep_bird_radiant/creep_bird_radiant_melee.vmdl', 1, 1, 350, 1, 20, 20, 150, 30, 1, 0, 50, 0, 6, 50),
]
for name, mdl, sc, ar, hp, hpR, amin, amax, atkRng, coin, faith, stat, exp, defPts, artType, artXP in _guard:
    UNITS[name] = mk('npc_dota_creature', mdl, sc, 160, 300, MG, AM,
        ar, hp, hpR, 0, 0, 0, amin, amax, 1.5, None, atkRng, HR,
        0, 500, CB, CD, coin, faith, stat, exp, defPts, artType, artXP)


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f'ERROR: Excel not found: {EXCEL_PATH}')
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['custom_units']

    # 1. 读取 Row2 表头 -> 列号映射
    field_to_col = {}
    for c in range(1, ws.max_column + 1):
        key = ws.cell(row=2, column=c).value
        if key:
            field_to_col[str(key).strip()] = c

    col_unitname = field_to_col.get('UnitName', 1)
    col_cnname = 2  # #LocUnitNameCn_{}

    # 2. 读取现有 UnitName -> Row 映射
    unit_rows = {}
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row=row, column=col_unitname).value
        if val:
            unit_rows[str(val).strip()] = row

    print(f'Excel 共 {len(unit_rows)} 个单位, {len(field_to_col)} 列')

    # 3. 重命名旧单位 (npc_enemy_zombie_lvl* -> npc_creep_train_tier*)
    for old_name, new_name in RENAME_MAP.items():
        if old_name in unit_rows:
            row = unit_rows[old_name]
            ws.cell(row=row, column=col_unitname).value = new_name
            unit_rows[new_name] = row
            del unit_rows[old_name]
            print(f'  重命名: {old_name} -> {new_name} (row {row})')

    # 4. 更新所有单位数据
    updated_count = 0
    for unit_name, fields in UNITS.items():
        row = unit_rows.get(unit_name)
        if not row:
            print(f'  WARNING: 找不到 {unit_name}, 跳过')
            continue

        # 更新中文名
        cn = CN_NAMES.get(unit_name)
        if cn:
            ws.cell(row=row, column=col_cnname).value = cn

        # 更新字段
        for field, value in fields.items():
            col = field_to_col.get(field)
            if not col:
                print(f'  WARNING: 找不到字段列 {field}')
                continue
            old_val = ws.cell(row=row, column=col).value
            if old_val != value:
                ws.cell(row=row, column=col).value = value
                updated_count += 1

    print(f'\n更新了 {updated_count} 个单元格')

    # 5. 保存
    wb.save(EXCEL_PATH)
    print(f'已保存到 {EXCEL_PATH}')
    print('请运行 yarn dev 重新生成 KV 文件')


if __name__ == '__main__':
    main()
