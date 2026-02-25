# -*- coding: utf-8 -*-
"""Update flame storm AbilityTextureName to use custom icon"""
import openpyxl

wb = openpyxl.load_workbook('技能表.xlsx')
ws = wb.active

flame_row = None
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v == 'ability_public_flame_storm':
        flame_row = r
        break

if flame_row is None:
    print('ERROR: flame storm not found!')
    wb.close()
    exit(1)

# Column 3 is usually AbilityTextureName
# Let's find which column has AbilityTextureName
header_row = 1
texture_col = None
for c in range(1, ws.max_column + 1):
    h = ws.cell(header_row, c).value
    if h and 'texturename' in str(h).lower().replace(' ', ''):
        texture_col = c
        break
    if h and 'texture' in str(h).lower():
        texture_col = c
        break

if texture_col:
    old_val = ws.cell(flame_row, texture_col).value
    ws.cell(flame_row, texture_col, 'flame_storm_icon')
    print(f'Updated AbilityTextureName col {texture_col}: {old_val} -> flame_storm_icon')
else:
    # Fallback: column 3 is common for AbilityTextureName
    for c in range(1, min(ws.max_column + 1, 50)):
        val = ws.cell(flame_row, c).value
        if val == 'lina_light_strike_array':
            ws.cell(flame_row, c, 'flame_storm_icon')
            print(f'Updated col {c}: lina_light_strike_array -> flame_storm_icon')
            break
    else:
        print('WARNING: Could not find lina_light_strike_array to replace')

wb.save('技能表.xlsx')
print('Done!')
wb.close()
