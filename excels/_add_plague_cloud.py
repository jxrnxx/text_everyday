# -*- coding: utf-8 -*-
"""Add ability_public_plague_cloud to 技能表.xlsx"""
import openpyxl

wb = openpyxl.load_workbook('技能表.xlsx')
ws = wb.active

# New row index
new_row = ws.max_row + 1  # Should be 9

# Based on the structure from martial_cleave (Row 8):
# Col 1:  技能名 = ability name
# Col 2:  基类 = BaseClass
# Col 3:  脚本路径 = ScriptFile
# Col 4:  图标 = AbilityTextureName
# Col 5:  技能行为 = AbilityBehavior
# Col 6:  最大等级 = MaxLevel
# Col 7:  冷却 = AbilityCooldown
# Col 8:  魔耗 = AbilityManaCost
# Col 11: 施法距离 = AbilityCastRange
# Col 12: 施法前摇 = AbilityCastPoint
# Col 15: 分类 = AbilityCategory
# Col 16: 星级 = AbilityStar
# Col 17: 元素 = AbilityElement
# Col 19-28: 数值1-10
# Col 31: 粒子 = particle precache
# Col 34: 名称(中)
# Col 35: 描述(中)

data = {
    1: 'ability_public_plague_cloud',
    2: 'ability_lua',
    3: 'abilities/ability_public_plague_cloud',
    4: 'venomancer_poison_nova',
    5: 'DOTA_ABILITY_BEHAVIOR_POINT | DOTA_ABILITY_BEHAVIOR_AOE',
    6: 4,  # MaxLevel
    7: 12,  # Cooldown
    8: 10,  # ManaCost
    11: 600,  # CastRange
    12: 0.3,  # CastPoint
    15: 'PUBLIC',
    16: 1,  # Star
    17: 'DIVINITY',
    19: 'radius 300',
    20: 'duration 5',
    21: 'dmg_multiplier 1 1.2 1.5 1.8',
    31: 'particles/units/heroes/hero_alchemist/alchemist_acid_spray.vpcf',
    34: '神念 · 瘟疫云',
    35: '主动：在目标区域降下毒云，每秒造成神念×倍率的法术伤害，持续5秒。',
}

for col, val in data.items():
    ws.cell(new_row, col, val)

wb.save('技能表.xlsx')
print(f'Done! Added plague cloud at row {new_row}')
