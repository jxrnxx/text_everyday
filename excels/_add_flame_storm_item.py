# -*- coding: utf-8 -*-
"""Fix flame storm item in 物品表.xlsx - correct the values that weren't overridden"""
import openpyxl

wb = openpyxl.load_workbook('物品表.xlsx')
ws = wb['npc_items_custom']

# Find flame storm row
flame_row = None
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 1).value == 'item_book_flame_storm_1':
        flame_row = r
        break

if flame_row is None:
    print('ERROR: flame storm not found!')
    wb.close()
    exit(1)

print('Found flame storm at row %d' % flame_row)

# Print all headers with their column number
headers = {}
for c in range(1, 30):
    h = ws.cell(1, c).value
    if h:
        headers[c] = str(h).strip()
        print('  C%d = %s  ->  current value: %s' % (c, h, ws.cell(flame_row, c).value))

# Now fix each column by matching header names
fixes = {
    'AbilityTextureName': 'item_book_flame_storm_1',
    'ID': 1404,
    'ItemQuality': 2,
    'LearnAbilityName': 'ability_public_flame_storm',
    'IconPath': 'file://{images}/spellicons/lina_light_strike_array.png',
    'SkillBookCategory': 'DIVINITY',
}

# Also check row 2 for header names (sometimes row 2 has English names)
row2_headers = {}
for c in range(1, 30):
    h = ws.cell(2, c).value
    if h:
        row2_headers[c] = str(h).strip()

print('')
print('Row 2 values (might be English headers):')
for c, h in row2_headers.items():
    print('  C%d = %s' % (c, h))

# Try to map using both row 1 and row 2 headers
for c in range(1, 30):
    h1 = headers.get(c, '')
    h2 = row2_headers.get(c, '')
    for fix_key, fix_val in fixes.items():
        if fix_key in h1 or fix_key in h2 or fix_key == h1 or fix_key == h2:
            old = ws.cell(flame_row, c).value
            ws.cell(flame_row, c, fix_val)
            print('Fixed C%d (%s): %s -> %s' % (c, h1 or h2, old, fix_val))

wb.save('物品表.xlsx')
print('')
print('Done! Saved fixes.')
wb.close()
