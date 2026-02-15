# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('技能表.xlsx')
ws = wb.active

lines = []
lines.append(f"Sheet: {ws.title}")
lines.append(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
lines.append("")

headers = []
for col in range(1, ws.max_column + 1):
    h = ws.cell(1, col).value
    headers.append(h if h else f"col_{col}")
    lines.append(f"  Col {col}: {h}")

lines.append("")

for row in range(2, ws.max_row + 1):
    name = ws.cell(row, 1).value
    if not name:
        continue
    lines.append(f"=== Row {row}: {name} ===")
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row, col).value
        if val is not None:
            lines.append(f"  [{col}] {headers[col-1]}: {val}")
    lines.append("")

with open('_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print("Done, wrote _output.txt")
