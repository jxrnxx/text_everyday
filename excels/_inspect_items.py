# -*- coding: utf-8 -*-
"""Compare golden_bell vs flame_storm in npc_items_custom sheet"""
import openpyxl

wb = openpyxl.load_workbook('物品表.xlsx')
ws = wb['npc_items_custom']

# Print per-column comparison
print('Col | Header | golden_bell (R16) | flame_storm (R17)')
print('-' * 80)
for c in range(1, 25):
    h = ws.cell(1, c).value or ''
    v16 = ws.cell(16, c).value or ''
    v17 = ws.cell(17, c).value or ''
    match = 'OK' if str(v16) == str(v17) or c in (1,) else ''
    print('C%-2d | %-20s | %-30s | %-30s %s' % (c, str(h)[:20], str(v16)[:30], str(v17)[:30], match))

wb.close()
