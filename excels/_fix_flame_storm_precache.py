# -*- coding: utf-8 -*-
"""Fix flame storm precache to use custom_flame_storm.vpcf"""
import openpyxl

wb = openpyxl.load_workbook('技能表.xlsx')
ws = wb.active

flame_row = None
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v == 'ability_public_flame_storm':
        flame_row = r
        break

if flame_row is None:
    print('ERROR: flame storm not found!')
    wb.close()
    exit(1)

ws.cell(flame_row, 31, 'particles/custom_flame_storm.vpcf')
print('Updated precache to: particles/custom_flame_storm.vpcf')

wb.save('技能表.xlsx')
print('Done!')
wb.close()
